import os
import sqlite3
import datetime
import pandas as pd

# Default database file path in the same directory as this script
DEFAULT_DB_PATH = os.path.abspath(os.path.join(os.path.dirname(__file__), "banquillo.db"))

def get_db_connection(db_path=DEFAULT_DB_PATH):
    """Establece conexión con la base de datos SQLite y activa soporte para llaves foráneas."""
    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA foreign_keys = ON;")
    conn.row_factory = sqlite3.Row
    return conn

def init_db(db_path=DEFAULT_DB_PATH, force_reset=False):
    """Inicializa la base de datos SQLite y crea las tablas necesarias."""
    conn = get_db_connection(db_path)
    cursor = conn.cursor()
    
    if force_reset:
        cursor.execute("DROP TABLE IF EXISTS citas;")
        cursor.execute("DROP TABLE IF EXISTS asignaciones;")
        cursor.execute("DROP TABLE IF EXISTS turnos_requeridos;")
        cursor.execute("DROP TABLE IF EXISTS disponibilidad;")
        cursor.execute("DROP TABLE IF EXISTS orientadores;")
        
    # Crear Tabla Orientadores
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS orientadores (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT NOT NULL UNIQUE,
        contacto TEXT,
        max_horas_semanales INTEGER DEFAULT 40
    );
    """)
    
    # Crear Tabla Disponibilidad
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS disponibilidad (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        orientador_id INTEGER NOT NULL,
        dia_semana TEXT,
        fecha TEXT NOT NULL,
        hora_inicio TEXT NOT NULL,
        hora_fin TEXT NOT NULL,
        barrio TEXT DEFAULT 'La Cumbre',
        FOREIGN KEY(orientador_id) REFERENCES orientadores(id) ON DELETE CASCADE,
        UNIQUE(orientador_id, fecha, hora_inicio, hora_fin, barrio)
    );
    """)
    
    # Crear Tabla Turnos Requeridos
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS turnos_requeridos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        dia_semana TEXT,
        fecha TEXT NOT NULL,
        hora_inicio TEXT NOT NULL,
        hora_fin TEXT NOT NULL,
        personas_requeridas INTEGER DEFAULT 1,
        barrio TEXT DEFAULT 'La Cumbre',
        UNIQUE(fecha, hora_inicio, hora_fin, barrio)
    );
    """)
    
    # Crear Tabla Asignaciones
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS asignaciones (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        turno_id INTEGER NOT NULL,
        orientador_id INTEGER NOT NULL,
        estado TEXT DEFAULT 'asignado',
        FOREIGN KEY(turno_id) REFERENCES turnos_requeridos(id) ON DELETE CASCADE,
        FOREIGN KEY(orientador_id) REFERENCES orientadores(id) ON DELETE CASCADE,
        UNIQUE(turno_id, orientador_id)
    );
    """)
    
    # Crear Tabla Citas
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS citas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        asignacion_id INTEGER NOT NULL,
        nombre_usuario TEXT,
        contacto_usuario TEXT,
        estado TEXT DEFAULT 'libre',
        FOREIGN KEY(asignacion_id) REFERENCES asignaciones(id) ON DELETE CASCADE,
        UNIQUE(asignacion_id)
    );
    """)
    
    conn.commit()
    conn.close()

# ----------------- FUNCIONES AUXILIARES DE NORMALIZACIÓN -----------------

def normalize_column_name(col):
    """Normaliza los nombres de columna del Excel para evitar problemas de codificación."""
    col = str(col).strip()
    col_lower = col.lower()
    if 'orientador' in col_lower:
        return 'orientador'
    if 'fecha' in col_lower:
        return 'fecha'
    if 'dia' in col_lower or 'da' in col_lower or 'día' in col_lower:
        return 'dia_semana'
    if 'inicio' in col_lower:
        return 'hora_inicio'
    if 'finali' in col_lower or 'fin' in col_lower:
        return 'hora_fin'
    if 'sesi' in col_lower or 'sesí' in col_lower:
        return 'numero_sesion'
    if 'usuario' in col_lower or 'nombre del' in col_lower:
        return 'nombre_usuario'
    if 'estado' in col_lower:
        return 'estado_cita'
    return col

def format_date(val):
    """Formatea valores de fecha al estándar YYYY-MM-DD."""
    if pd.isna(val):
        return None
    if isinstance(val, pd.Timestamp):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, (datetime.date, datetime.datetime)):
        return val.strftime('%Y-%m-%d')
    try:
        dt = pd.to_datetime(val)
        return dt.strftime('%Y-%m-%d')
    except Exception:
        return str(val).strip()

def format_time(val):
    """Formatea valores de hora al estándar HH:MM:SS."""
    if pd.isna(val):
        return None
    if isinstance(val, datetime.time):
        return val.strftime('%H:%M:%S')
    if isinstance(val, pd.Timestamp):
        return val.strftime('%H:%M:%S')
    if isinstance(val, datetime.datetime):
        return val.time().strftime('%H:%M:%S')
    
    s = str(val).strip()
    parts = s.split(':')
    if len(parts) == 2:
        return f"{parts[0].zfill(2)}:{parts[1].zfill(2)}:00"
    elif len(parts) == 3:
        return f"{parts[0].zfill(2)}:{parts[1].zfill(2)}:{parts[2].zfill(2)}"
    return s

def get_sheet_name(year, month):
    """Retorna el nombre correspondiente de la hoja de Excel basado en año y mes."""
    months_es = {
        1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
        5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
        9: "Septiembre", 10: "Octubre", 11: "Noviembre",
        12: "Diciembre"
    }
    month_name = months_es.get(month, f"Mes {month}")
    if year == 2025 and month == 12:
        return "Diciembre 20"
    return f"{month_name}"

# ----------------- CRUD ORIENTADORES -----------------

def create_orientador(nombre, contacto=None, max_horas_semanales=40, db_path=DEFAULT_DB_PATH):
    conn = get_db_connection(db_path)
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO orientadores (nombre, contacto, max_horas_semanales) VALUES (?, ?, ?)",
        (nombre.strip(), contacto, max_horas_semanales)
    )
    new_id = cursor.lastrowid
    conn.commit()
    conn.close()
    return new_id

def get_orientadores(db_path=DEFAULT_DB_PATH):
    conn = get_db_connection(db_path)
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM orientadores ORDER BY nombre ASC")
    rows = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return rows

def get_orientador_by_id(orientador_id, db_path=DEFAULT_DB_PATH):
    conn = get_db_connection(db_path)
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM orientadores WHERE id = ?", (orientador_id,))
    row = cursor.fetchone()
    conn.close()
    return dict(row) if row else None

def get_orientador_by_name(nombre, db_path=DEFAULT_DB_PATH):
    conn = get_db_connection(db_path)
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM orientadores WHERE nombre = ?", (nombre.strip(),))
    row = cursor.fetchone()
    conn.close()
    return dict(row) if row else None

def update_orientador(orientador_id, nombre=None, contacto=None, max_horas_semanales=None, db_path=DEFAULT_DB_PATH):
    conn = get_db_connection(db_path)
    cursor = conn.cursor()
    
    fields = []
    params = []
    if nombre is not None:
        fields.append("nombre = ?")
        params.append(nombre.strip())
    if contacto is not None:
        fields.append("contacto = ?")
        params.append(contacto)
    if max_horas_semanales is not None:
        fields.append("max_horas_semanales = ?")
        params.append(max_horas_semanales)
        
    if fields:
        params.append(orientador_id)
        query = f"UPDATE orientadores SET {', '.join(fields)} WHERE id = ?"
        cursor.execute(query, params)
        conn.commit()
    conn.close()

def delete_orientador(orientador_id, db_path=DEFAULT_DB_PATH):
    conn = get_db_connection(db_path)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM orientadores WHERE id = ?", (orientador_id,))
    conn.commit()
    conn.close()

# ----------------- CRUD DISPONIBILIDAD -----------------

def create_disponibilidad(orientador_id, dia_semana, fecha, hora_inicio, hora_fin, barrio='La Cumbre', db_path=DEFAULT_DB_PATH):
    conn = get_db_connection(db_path)
    cursor = conn.cursor()
    fecha_fmt = format_date(fecha)
    hora_ini_fmt = format_time(hora_inicio)
    hora_fin_fmt = format_time(hora_fin)
    
    cursor.execute(
        """INSERT INTO disponibilidad (orientador_id, dia_semana, fecha, hora_inicio, hora_fin, barrio) 
           VALUES (?, ?, ?, ?, ?, ?)""",
        (orientador_id, dia_semana, fecha_fmt, hora_ini_fmt, hora_fin_fmt, barrio)
    )
    new_id = cursor.lastrowid
    conn.commit()
    conn.close()
    return new_id

def get_disponibilidades(db_path=DEFAULT_DB_PATH):
    conn = get_db_connection(db_path)
    cursor = conn.cursor()
    cursor.execute(
        """SELECT d.*, o.nombre AS orientador_nombre 
           FROM disponibilidad d 
           JOIN orientadores o ON d.orientador_id = o.id 
           ORDER BY d.fecha ASC, d.hora_inicio ASC"""
    )
    rows = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return rows

def get_disponibilidad_by_orientador(orientador_id, db_path=DEFAULT_DB_PATH):
    conn = get_db_connection(db_path)
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM disponibilidad WHERE orientador_id = ? ORDER BY fecha ASC, hora_inicio ASC", (orientador_id,))
    rows = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return rows

def delete_disponibilidad(disponibilidad_id, db_path=DEFAULT_DB_PATH):
    conn = get_db_connection(db_path)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM disponibilidad WHERE id = ?", (disponibilidad_id,))
    conn.commit()
    conn.close()

def get_disponibilidades_by_barrio(barrio, db_path=DEFAULT_DB_PATH):
    conn = get_db_connection(db_path)
    cursor = conn.cursor()
    cursor.execute(
        """SELECT d.*, o.nombre AS orientador_nombre 
           FROM disponibilidad d 
           JOIN orientadores o ON d.orientador_id = o.id 
           WHERE d.barrio = ?
           ORDER BY d.fecha ASC, d.hora_inicio ASC""",
        (barrio,)
    )
    rows = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return rows

def get_turnos_requeridos_by_barrio(barrio, db_path=DEFAULT_DB_PATH):
    conn = get_db_connection(db_path)
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM turnos_requeridos WHERE barrio = ? ORDER BY fecha ASC, hora_inicio ASC", (barrio,))
    rows = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return rows

# ----------------- CRUD TURNOS REQUERIDOS -----------------

def create_turno_requerido(dia_semana, fecha, hora_inicio, hora_fin, personas_requeridas=1, barrio='La Cumbre', db_path=DEFAULT_DB_PATH):
    conn = get_db_connection(db_path)
    cursor = conn.cursor()
    fecha_fmt = format_date(fecha)
    hora_ini_fmt = format_time(hora_inicio)
    hora_fin_fmt = format_time(hora_fin)
    
    cursor.execute(
        """INSERT INTO turnos_requeridos (dia_semana, fecha, hora_inicio, hora_fin, personas_requeridas, barrio) 
           VALUES (?, ?, ?, ?, ?, ?)""",
        (dia_semana, fecha_fmt, hora_ini_fmt, hora_fin_fmt, personas_requeridas, barrio)
    )
    new_id = cursor.lastrowid
    conn.commit()
    conn.close()
    return new_id

def get_turnos_requeridos(db_path=DEFAULT_DB_PATH):
    conn = get_db_connection(db_path)
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM turnos_requeridos ORDER BY fecha ASC, hora_inicio ASC")
    rows = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return rows

def get_turno_requerido_by_id(turno_id, db_path=DEFAULT_DB_PATH):
    conn = get_db_connection(db_path)
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM turnos_requeridos WHERE id = ?", (turno_id,))
    row = cursor.fetchone()
    conn.close()
    return dict(row) if row else None

def update_turno_requerido(turno_id, dia_semana=None, fecha=None, hora_inicio=None, hora_fin=None, personas_requeridas=None, barrio=None, db_path=DEFAULT_DB_PATH):
    conn = get_db_connection(db_path)
    cursor = conn.cursor()
    
    fields = []
    params = []
    if dia_semana is not None:
        fields.append("dia_semana = ?")
        params.append(dia_semana)
    if fecha is not None:
        fields.append("fecha = ?")
        params.append(format_date(fecha))
    if hora_inicio is not None:
        fields.append("hora_inicio = ?")
        params.append(format_time(hora_inicio))
    if hora_fin is not None:
        fields.append("hora_fin = ?")
        params.append(format_time(hora_fin))
    if personas_requeridas is not None:
        fields.append("personas_requeridas = ?")
        params.append(personas_requeridas)
    if barrio is not None:
        fields.append("barrio = ?")
        params.append(barrio)
        
    if fields:
        params.append(turno_id)
        query = f"UPDATE turnos_requeridos SET {', '.join(fields)} WHERE id = ?"
        cursor.execute(query, params)
        conn.commit()
    conn.close()

def delete_turno_requerido(turno_id, db_path=DEFAULT_DB_PATH):
    conn = get_db_connection(db_path)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM turnos_requeridos WHERE id = ?", (turno_id,))
    conn.commit()
    conn.close()

# ----------------- CRUD ASIGNACIONES -----------------

def create_asignacion(turno_id, orientador_id, estado='asignado', db_path=DEFAULT_DB_PATH):
    conn = get_db_connection(db_path)
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO asignaciones (turno_id, orientador_id, estado) VALUES (?, ?, ?)",
        (turno_id, orientador_id, estado)
    )
    new_id = cursor.lastrowid
    conn.commit()
    conn.close()
    return new_id

def get_asignaciones(db_path=DEFAULT_DB_PATH):
    conn = get_db_connection(db_path)
    cursor = conn.cursor()
    cursor.execute(
        """SELECT a.*, o.nombre AS orientador_nombre, 
                  tr.fecha, tr.hora_inicio, tr.hora_fin, tr.dia_semana 
           FROM asignaciones a 
           JOIN orientadores o ON a.orientador_id = o.id 
           JOIN turnos_requeridos tr ON a.turno_id = tr.id 
           ORDER BY tr.fecha ASC, tr.hora_inicio ASC"""
    )
    rows = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return rows

def get_asignacion_by_id(asignacion_id, db_path=DEFAULT_DB_PATH):
    conn = get_db_connection(db_path)
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM asignaciones WHERE id = ?", (asignacion_id,))
    row = cursor.fetchone()
    conn.close()
    return dict(row) if row else None

def update_asignacion_status(asignacion_id, estado, db_path=DEFAULT_DB_PATH):
    conn = get_db_connection(db_path)
    cursor = conn.cursor()
    cursor.execute("UPDATE asignaciones SET estado = ? WHERE id = ?", (estado, asignacion_id))
    conn.commit()
    conn.close()

def delete_asignacion(asignacion_id, db_path=DEFAULT_DB_PATH):
    conn = get_db_connection(db_path)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM asignaciones WHERE id = ?", (asignacion_id,))
    conn.commit()
    conn.close()

# ----------------- CRUD CITAS -----------------

def create_cita(asignacion_id, nombre_usuario=None, contacto_usuario=None, estado='libre', db_path=DEFAULT_DB_PATH):
    conn = get_db_connection(db_path)
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO citas (asignacion_id, nombre_usuario, contacto_usuario, estado) VALUES (?, ?, ?, ?)",
        (asignacion_id, nombre_usuario, contacto_usuario, estado)
    )
    new_id = cursor.lastrowid
    conn.commit()
    conn.close()
    return new_id

def get_citas(db_path=DEFAULT_DB_PATH):
    conn = get_db_connection(db_path)
    cursor = conn.cursor()
    cursor.execute(
        """SELECT c.*, o.nombre AS orientador_nombre, 
                  tr.fecha, tr.hora_inicio, tr.hora_fin 
           FROM citas c 
           JOIN asignaciones a ON c.asignacion_id = a.id 
           JOIN orientadores o ON a.orientador_id = o.id 
           JOIN turnos_requeridos tr ON a.turno_id = tr.id 
           ORDER BY tr.fecha ASC, tr.hora_inicio ASC"""
    )
    rows = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return rows

def get_cita_by_id(cita_id, db_path=DEFAULT_DB_PATH):
    conn = get_db_connection(db_path)
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM citas WHERE id = ?", (cita_id,))
    row = cursor.fetchone()
    conn.close()
    return dict(row) if row else None

def update_cita(cita_id, nombre_usuario=None, contacto_usuario=None, estado=None, db_path=DEFAULT_DB_PATH):
    conn = get_db_connection(db_path)
    cursor = conn.cursor()
    
    fields = []
    params = []
    if nombre_usuario is not None:
        fields.append("nombre_usuario = ?")
        params.append(nombre_usuario)
    if contacto_usuario is not None:
        fields.append("contacto_usuario = ?")
        params.append(contacto_usuario)
    if estado is not None:
        fields.append("estado = ?")
        params.append(estado)
        
    if fields:
        params.append(cita_id)
        query = f"UPDATE citas SET {', '.join(fields)} WHERE id = ?"
        cursor.execute(query, params)
        conn.commit()
    conn.close()

def delete_cita(cita_id, db_path=DEFAULT_DB_PATH):
    conn = get_db_connection(db_path)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM citas WHERE id = ?", (cita_id,))
    conn.commit()
    conn.close()

# ----------------- IMPORTACIÓN / EXPORTACIÓN EXCEL -----------------

def import_from_excel(excel_path, db_path=DEFAULT_DB_PATH, reset=True):
    """Importa los orientadores, disponibilidades, turnos y asignaciones desde un archivo Excel."""
    # 1. Inicializar la base de datos (con opción de limpieza completa)
    init_db(db_path, force_reset=reset)
    
    # 2. Cargar Orientadores
    df_o = pd.read_excel(excel_path, sheet_name="Orientadores")
    df_o.columns = ['nombre', 'contacto']
    df_o['nombre'] = df_o['nombre'].astype(str).str.strip()
    df_o['contacto'] = df_o['contacto'].apply(lambda x: str(x).strip() if pd.notna(x) else None)
    
    conn = get_db_connection(db_path)
    cursor = conn.cursor()
    
    orientadores_map = {} # nombre -> id
    for _, row in df_o.iterrows():
        nombre = row['nombre']
        if not nombre or nombre.lower() == 'nan':
            continue
        contacto = row['contacto']
        cursor.execute(
            "INSERT OR IGNORE INTO orientadores (nombre, contacto, max_horas_semanales) VALUES (?, ?, ?)",
            (nombre, contacto, 40)
        )
        cursor.execute("SELECT id FROM orientadores WHERE nombre = ?", (nombre,))
        orientadores_map[nombre] = cursor.fetchone()[0]
        
    # 3. Leer hojas de agendas
    xls = pd.ExcelFile(excel_path)
    all_schedules = []
    
    for sheet in xls.sheet_names:
        if sheet in ['Orientadores', 'Hoja 1']:
            continue
            
        df = pd.read_excel(excel_path, sheet_name=sheet)
        df.columns = [normalize_column_name(c) for c in df.columns]
        
        # Verificar si es una hoja de agenda válida
        required_cols = {'orientador', 'fecha', 'hora_inicio', 'hora_fin'}
        if not required_cols.issubset(df.columns):
            continue
            
        for _, row in df.iterrows():
            orientador_name = str(row['orientador']).strip() if pd.notna(row['orientador']) else None
            if not orientador_name or orientador_name.lower() == 'nan':
                continue
                
            fecha_val = format_date(row['fecha'])
            hora_ini_val = format_time(row['hora_inicio'])
            hora_fin_val = format_time(row['hora_fin'])
            dia_semana_val = str(row['dia_semana']).strip() if 'dia_semana' in row and pd.notna(row['dia_semana']) else None
            
            nombre_usuario_val = str(row['nombre_usuario']).strip() if 'nombre_usuario' in row and pd.notna(row['nombre_usuario']) else None
            estado_cita_val = str(row['estado_cita']).strip() if 'estado_cita' in row and pd.notna(row['estado_cita']) else None
            
            if not (fecha_val and hora_ini_val and hora_fin_val):
                continue
                
            all_schedules.append({
                'orientador': orientador_name,
                'dia_semana': dia_semana_val,
                'fecha': fecha_val,
                'hora_inicio': hora_ini_val,
                'hora_fin': hora_fin_val,
                'nombre_usuario': nombre_usuario_val,
                'estado_cita': estado_cita_val
            })
            
    # Eliminar duplicados de asignación entre hojas redundantes (como Hoja 4)
    unique_schedules = []
    seen = set()
    for item in all_schedules:
        key = (item['orientador'], item['fecha'], item['hora_inicio'], item['hora_fin'])
        if key not in seen:
            seen.add(key)
            unique_schedules.append(item)
            
    # Garantizar que todos los orientadores existan en la BD
    for item in unique_schedules:
        name = item['orientador']
        if name not in orientadores_map:
            cursor.execute("INSERT OR IGNORE INTO orientadores (nombre, contacto, max_horas_semanales) VALUES (?, ?, ?)", (name, None, 40))
            cursor.execute("SELECT id FROM orientadores WHERE nombre = ?", (name,))
            orientadores_map[name] = cursor.fetchone()[0]
            
    # Insertar disponibilidades
    for item in unique_schedules:
        o_id = orientadores_map[item['orientador']]
        cursor.execute(
            """INSERT OR IGNORE INTO disponibilidad 
               (orientador_id, dia_semana, fecha, hora_inicio, hora_fin, barrio) 
               VALUES (?, ?, ?, ?, ?, ?)""",
            (o_id, item['dia_semana'], item['fecha'], item['hora_inicio'], item['hora_fin'], 'La Cumbre')
        )
        
    # Agrupar por franja horaria para crear los turnos requeridos y contar personas requeridas
    turnos_group = {}
    for item in unique_schedules:
        key = (item['fecha'], item['hora_inicio'], item['hora_fin'])
        if key not in turnos_group:
            turnos_group[key] = {
                'dia_semana': item['dia_semana'],
                'orientadores': set(),
                'rows': []
            }
        turnos_group[key]['orientadores'].add(item['orientador'])
        turnos_group[key]['rows'].append(item)
        
    # Registrar turnos, asignaciones y citas
    for (fecha, hora_ini, hora_fin), info in turnos_group.items():
        personas_req = len(info['orientadores'])
        
        # Buscar o insertar turno
        cursor.execute(
            "SELECT id FROM turnos_requeridos WHERE fecha = ? AND hora_inicio = ? AND hora_fin = ? AND barrio = ?",
            (fecha, hora_ini, hora_fin, 'La Cumbre')
        )
        res = cursor.fetchone()
        if res:
            turno_id = res[0]
            cursor.execute(
                "UPDATE turnos_requeridos SET personas_requeridas = ? WHERE id = ?",
                (personas_req, turno_id)
            )
        else:
            cursor.execute(
                """INSERT INTO turnos_requeridos 
                   (dia_semana, fecha, hora_inicio, hora_fin, personas_requeridas, barrio) 
                   VALUES (?, ?, ?, ?, ?, ?)""",
                (info['dia_semana'], fecha, hora_ini, hora_fin, personas_req, 'La Cumbre')
            )
            turno_id = cursor.lastrowid
            
        # Registrar asignaciones y citas
        for row in info['rows']:
            o_id = orientadores_map[row['orientador']]
            
            # Buscar o insertar asignación
            cursor.execute(
                "SELECT id FROM asignaciones WHERE turno_id = ? AND orientador_id = ?",
                (turno_id, o_id)
            )
            res_asig = cursor.fetchone()
            if res_asig:
                asignacion_id = res_asig[0]
            else:
                cursor.execute(
                    "INSERT INTO asignaciones (turno_id, orientador_id, estado) VALUES (?, ?, ?)",
                    (turno_id, o_id, 'asignado')
                )
                asignacion_id = cursor.lastrowid
                
            # Buscar o insertar cita
            cursor.execute("SELECT id FROM citas WHERE asignacion_id = ?", (asignacion_id,))
            res_cita = cursor.fetchone()
            
            nombre_u = row['nombre_usuario']
            estado_c = row['estado_cita']
            if not nombre_u or nombre_u.lower() == 'nan':
                nombre_u = None
                contacto_u = None
                estado_c = 'libre'
            else:
                contacto_u = None
                if not estado_c or estado_c.lower() == 'nan':
                    estado_c = 'reservada'
                    
            if res_cita:
                cursor.execute(
                    """UPDATE citas 
                       SET nombre_usuario = ?, contacto_usuario = ?, estado = ? 
                       WHERE asignacion_id = ?""",
                    (nombre_u, contacto_u, estado_c, asignacion_id)
                )
            else:
                cursor.execute(
                    """INSERT INTO citas (asignacion_id, nombre_usuario, contacto_usuario, estado) 
                       VALUES (?, ?, ?, ?)""",
                    (asignacion_id, nombre_u, contacto_u, estado_c)
                )
                
    conn.commit()
    conn.close()
    print(f"Migración completada. Registrados: {len(orientadores_map)} orientadores, {len(turnos_group)} turnos únicos.")

def export_to_excel(excel_path, db_path=DEFAULT_DB_PATH):
    """Exporta los datos de la base de datos a un archivo Excel formateado de forma similar al original."""
    conn = get_db_connection(db_path)
    
    # 1. Pestaña Orientadores
    df_o = pd.read_sql_query(
        "SELECT nombre AS [Unnamed: 0], contacto AS [Número de contacto] FROM orientadores ORDER BY nombre ASC",
        conn
    )
    
    # 2. Obtener todas las asignaciones y citas
    query = """
        SELECT 
            c.id AS [Número de sesión], 
            o.nombre AS [Orientador(a)], 
            tr.dia_semana AS [Día de la semana], 
            tr.fecha AS [Fecha], 
            tr.hora_inicio AS [Hora inicio], 
            tr.hora_fin AS [Hora finalización], 
            c.nombre_usuario AS [Nombre del usuario], 
            c.estado AS [Estado de la cita],
            tr.barrio AS [Barrio]
        FROM asignaciones a
        JOIN orientadores o ON a.orientador_id = o.id
        JOIN turnos_requeridos tr ON a.turno_id = tr.id
        LEFT JOIN citas c ON c.asignacion_id = a.id
        ORDER BY tr.fecha ASC, tr.hora_inicio ASC, o.nombre ASC
    """
    df_all = pd.read_sql_query(query, conn)
    conn.close()
    
    # Convertir Fecha a datetime para agrupamiento
    df_all['datetime_fecha'] = pd.to_datetime(df_all['Fecha'])
    
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        # Escribir pestaña de Orientadores
        df_o.to_excel(writer, sheet_name="Orientadores", index=False)
        
        # Agrupar por Año y Mes
        grouped = df_all.groupby([df_all['datetime_fecha'].dt.year, df_all['datetime_fecha'].dt.month])
        for (year, month), group in grouped:
            sheet_name = get_sheet_name(year, month)
            
            # Limpiar columnas temporales y ordenar
            group_export = group.drop(columns=['datetime_fecha'])
            
            # Si el número de sesión está vacío o es nulo en las filas 'libre', representarlo como NaN/vacío
            group_export.loc[group_export['Estado de la cita'] == 'libre', 'Número de sesión'] = None
            
            group_export.to_excel(writer, sheet_name=sheet_name, index=False)
            
    # Darle formato estético y profesional al Excel generado
    _format_excel_styling(excel_path)
    print(f"Exportación exitosa a: {excel_path}")

def _format_excel_styling(file_path):
    """Aplica formato estético a las columnas y encabezados del Excel generado."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.load_workbook(file_path)
    
    # Estilos
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid') # Azul corporativo
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Formatear la fila de encabezados
        ws.row_dimensions[1].height = 25
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            
        # Formatear filas de datos y ajustar anchos
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            col_name = str(col[0].value).lower()
            
            for cell in col:
                val_str = str(cell.value or '')
                max_len = max(max_len, len(val_str))
                
                # Alinear datos adecuadamente
                if cell.row > 1:
                    if any(k in col_name for k in ['fecha', 'hora', 'dia', 'sesi', 'contacto', 'estado']):
                        cell.alignment = center_align
                    else:
                        cell.alignment = left_align
                        
            # Ajustar ancho de columna con margen
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
    wb.save(file_path)

if __name__ == '__main__':
    # Si se ejecuta directamente, realiza la inicialización e importación inicial
    print("Iniciando db_manager...")
    excel_source = os.path.abspath(os.path.join(os.path.dirname(__file__), "AGENDAS - LA CUMBRE.xlsx"))
    if os.path.exists(excel_source):
        print(f"Detectado archivo origen: {excel_source}")
        import_from_excel(excel_source)
    else:
        print(f"No se detectó el archivo de Excel origen en la ruta por defecto: {excel_source}")
        init_db()
